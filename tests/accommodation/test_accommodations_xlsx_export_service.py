from __future__ import annotations
from io import BytesIO

import pytest

from accommodation.services.accommodations_xlsx_export_service import HEADERS, AccommodationExportRow
from accommodation.services.accommodations_xlsx_export_service import build_accommodation_export_rows
from accommodation.services.accommodations_xlsx_export_service import build_postal_code_geo_index
from accommodation.services.accommodations_xlsx_export_service import compute_total_availability
from accommodation.services.accommodations_xlsx_export_service import department_code_from_postal_code
from accommodation.services.accommodations_xlsx_export_service import export_accommodations_to_xlsx
from accommodation.services.accommodations_xlsx_export_service import resolve_department_and_academy

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - runtime dependency guard
    raise ImportError("openpyxl is required for accommodations XLSX export tests") from exc


@pytest.fixture(autouse=True)
def create_owners_group():
    # Override global DB fixture from conftest for pure unit tests in this module.
    return None


@pytest.mark.parametrize(
    ("postal_code", "expected"),
    [
        ("75011", "75"),
        ("97100", "971"),
        ("98845", "988"),
        ("", ""),
        ("abcde", ""),
    ],
)
def test_department_code_from_postal_code(postal_code, expected):
    assert department_code_from_postal_code(postal_code) == expected


def test_build_postal_code_geo_index_keeps_first_match():
    city_rows = [
        (["75011"], "Paris", "Ile-de-France"),
        (["75011", "75012"], "Other Department", "Other Region"),
    ]

    postal_code_to_geo = build_postal_code_geo_index(city_rows)

    assert postal_code_to_geo["75011"] == ("Paris", "Ile-de-France")
    assert postal_code_to_geo["75012"] == ("Other Department", "Other Region")


def test_resolve_department_and_academy_with_fallback():
    postal_code_to_geo = {"75011": ("Paris", "Ile-de-France")}
    departments_by_code = {"69": ("Rhone", "Auvergne-Rhone-Alpes")}

    assert resolve_department_and_academy("75011", postal_code_to_geo, departments_by_code) == (
        "Paris",
        "Ile-de-France",
    )
    assert resolve_department_and_academy("69002", postal_code_to_geo, departments_by_code) == (
        "Rhone",
        "Auvergne-Rhone-Alpes",
    )
    assert resolve_department_and_academy("abc", postal_code_to_geo, departments_by_code) == ("", "")


def test_build_accommodation_export_rows():
    accommodation_rows = [
        (
            "Residence A",
            "Owner A",
            120,
            "75011",
            1,
            2,
            3,
            4,
            5,
            6,
            7,
            8,
            "Résidence Universitaire conventionnée",
            100000,
            150000,
        ),
        (
            "Residence B",
            None,
            None,
            "69002",
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            "Résidence Hôtelière à vocation sociale",
            100000,
            150000,
        ),
    ]
    postal_code_to_geo = {"75011": ("Paris", "Ile-de-France")}
    departments_by_code = {"69": ("Rhone", "Auvergne-Rhone-Alpes")}

    rows = build_accommodation_export_rows(accommodation_rows, postal_code_to_geo, departments_by_code)

    assert rows == [
        AccommodationExportRow(
            name="Residence A",
            owner_name="Owner A",
            nb_total_apartments=120,
            postal_code="75011",
            department_name="Paris",
            academy_name="Ile-de-France",
            has_availability=True,
            residence_type="Résidence Universitaire conventionnée",
            price_min=100000,
            price_max=150000,
        ),
        AccommodationExportRow(
            name="Residence B",
            owner_name="",
            nb_total_apartments=None,
            postal_code="69002",
            department_name="Rhone",
            academy_name="Auvergne-Rhone-Alpes",
            has_availability=False,
            residence_type="Résidence Hôtelière à vocation sociale",
            price_min=100000,
            price_max=150000,
        ),
    ]


def test_compute_total_availability():
    assert compute_total_availability([None, None, None]) is None
    assert compute_total_availability([1, None, 2, 0]) == 3


def test_export_accommodations_to_xlsx():
    rows = [
        AccommodationExportRow(
            name="Residence A",
            owner_name="Owner A",
            nb_total_apartments=100,
            postal_code="75011",
            department_name="Paris",
            academy_name="Ile-de-France",
            has_availability=True,
            residence_type="Résidence Universitaire conventionnée",
            price_min=100000,
            price_max=150000,
        ),
        AccommodationExportRow(
            name="Residence B",
            owner_name="Owner B",
            nb_total_apartments=80,
            postal_code="69002",
            department_name="Rhone",
            academy_name="Auvergne-Rhone-Alpes",
            has_availability=False,
            residence_type="Résidence Hôtelière à vocation sociale",
            price_min=100000,
            price_max=150000,
        ),
    ]

    buffer, exported_count = export_accommodations_to_xlsx(rows)

    assert exported_count == 2
    assert isinstance(buffer, BytesIO)

    buffer.seek(0)
    content = buffer.getvalue()
    assert len(content) > 0

    workbook = load_workbook(filename=buffer)
    worksheet = workbook.active

    assert list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0] == tuple(HEADERS)
    assert list(worksheet.iter_rows(min_row=2, max_row=2, values_only=True))[0] == tuple(rows[0].as_xlsx_row())
    assert list(worksheet.iter_rows(min_row=3, max_row=3, values_only=True))[0] == tuple(rows[1].as_xlsx_row())
